# Imports
import openpyxl.workbook
from front import Ui_MainWindow
from PyQt5 import QtWidgets
import sys
from PyQt5.QtWidgets import QMessageBox, QFileDialog
from PyQt5.QtCore import pyqtSignal, QObject, QThread
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os
import sys
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options

# Global Variables
global dir_name                        # Global variable to hold the selected directory name.
global number_of_files                 # Global variable that contains the number of html files in the directory.
number_of_files = 0                    # Initalizing the number of files global variable to 0 files
global progress_bar_counter            # Global variable holding the current progress bar counter
progress_bar_counter = 0               # Initalizing the progress bar counter to 0
global progress_bar_step               # Global variable for the step of the progress bar based on the number of files to process
global files_list                      # Global list of files that contains the file names of the selected directory.
options = Options()                    # Initalizing options variable to contain chrome driver options.
options.add_argument('--headless=new') # Headless option to run chrome in silent mode.
driver = webdriver.Chrome(options)     # Initalizing the WebDriver to google chrome with the created options.
test_results_excel_file = "Tests_Results.xlsx" # Global Variable holding the output excel file name


# Functions
def check_if_excel_exists(excel_file_name):
    """
    Checks if an Excel file exists, and returns an openpyxl Workbook object.

    If the specified Excel file exists, it loads and returns the workbook. 
    If the file does not exist, it creates a new workbook, saves it with the 
    specified filename, and returns the new workbook.

    Args:
        excel_file_name (str): The name or path of the Excel file to check.

    Returns:
        openpyxl.Workbook: The loaded or newly created workbook object.
    """
    if os.path.exists(excel_file_name):
        workbook = openpyxl.load_workbook(excel_file_name)
    else:
        workbook = openpyxl.Workbook()
        workbook.save(excel_file_name)
    return workbook


def is_sheet_empty(excel_file):
  """Checks if a specified sheet in an Excel file is empty.

  Args:
    excel_file: Path to the Excel file.

  Returns:
    True if the sheet is empty, False otherwise.
  """
  workbook = openpyxl.load_workbook(excel_file)
  sheet = workbook.active
  # Check if the sheet has any rows
  if sheet.max_row == 1 and sheet['A1'].value is None:
    return True
  else:
    return False


def write_data_excel(excel_file_name, excel_data, styling_options):
    """
    Writes data to an Excel file with specified styling options and auto-fits the columns and rows.

    This function writes a list of data to an Excel file. If the file doesn't exist, it creates a new one. 
    The function also applies specified styling options to the cells and adjusts the column widths and row heights 
    to fit the content.

    Args:
        excel_file_name (str): The name or path of the Excel file to write data to.
        excel_data (list): A list of lists, where each sublist represents a row of data to be written to the Excel file.
        styling_options (dict): A dictionary of styling options. The default options are:
            - 'font': openpyxl.styles.Font object (default: Arial, size 14, not bold)
            - 'alignment': openpyxl.styles.Alignment object (default: centered horizontally and vertically)

    Returns:
        None
    """
    # Defining the default styling options
    default_options = {
        "font": Font(name='Arial', size=14, bold=False),
        "alignment": Alignment(horizontal='center', vertical='center')
    }
    # check if there are any styling options passed to the function and combine them with default options
    styling_options = {**default_options, **styling_options}
    workbook = check_if_excel_exists(excel_file_name)   
    # Load the workbook and select the active sheet
    sheet = workbook.active
    if is_sheet_empty(excel_file_name):
        next_row = sheet.max_row
    else:
        next_row = sheet.max_row + 1
    for row_data in excel_data:
        for col, data in enumerate(row_data, start=1):
            cell = sheet.cell(row=next_row, column=col, value=data)
            cell.font = styling_options.get('font')
            cell.alignment = styling_options.get('alignment')
        next_row += 1
    
    # Autofitting the rows and columns the adjusted width and height formulas are obtained by trial and error and not accurate
    # Autofit columns
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
                    font_size = styling_options.get('font').size
    
            except:
                pass
        adjusted_width = max_length * font_size / 12 + 2
        sheet.column_dimensions[column].width = adjusted_width
    # Autofit rows
    for row in sheet.rows:
        max_height = 0
        for cell in row:
            try:
                if len(str(cell.value)) > max_height:
                    max_height = len(str(cell.value))
            except:
                pass
        adjusted_height = max_height * font_size / 12 + 2
        sheet.row_dimensions[row[0].row].height = adjusted_height
    
    workbook.save(excel_file_name)

def format_cells_with_values(excel_file_name, search_value, font=None, fill=None, alignment=None):
    """
    Formats cells in an Excel file that contain a specific search value with given styling options.

    This function loads an Excel file and iterates through all cells in the active sheet. 
    If a cell contains the specified search value, it applies the provided formatting options 
    (font, fill, and alignment) to that cell.

    Args:
        excel_file_name (str): The name or path of the Excel file to be formatted.
        search_value (any): The value to search for in the cells. Cells containing this value will be formatted.
        font (openpyxl.styles.Font, optional): The font to apply to the cells containing the search value. Default is None.
        fill (openpyxl.styles.PatternFill, optional): The fill pattern to apply to the cells containing the search value. Default is None.
        alignment (openpyxl.styles.Alignment, optional): The alignment to apply to the cells containing the search value. Default is None.

    Returns:
        None
    """
    # Load the workbook and select the active sheet
    workbook = openpyxl.load_workbook(excel_file_name)
    sheet = workbook.active
    
    # Iterate over all rows and columns in the sheet
    for row in sheet.iter_rows():
        for cell in row:
            # Check if the cell value matches the search value
            if cell.value == search_value:
                # Apply formatting if value matches
                if font:
                    cell.font = font
                if fill:
                    cell.fill = fill
                if alignment:
                    cell.alignment = alignment
    
    # Save the workbook
    workbook.save(excel_file_name)


# Classes
class Worker(QObject):
    # Signals to control the GUI using threading
    progress_updated = pyqtSignal(int)  # Signal to update progress
    status_update = pyqtSignal(str)     # Signal to send status updates
    task_completed = pyqtSignal()       # Signal of task completion
    task_error = pyqtSignal()           # Signal of error

    def __init__(self, parent=None):
        super().__init__()
        self.parent_widget = parent

    def run(self):
        try:
            # Remove any previous excel files to start with a new one
            if os.path.exists(test_results_excel_file):
                os.remove(test_results_excel_file)
            # Variable to check if this is the first iteration.
            first_iteration = True
            # Check if there are HTML files in the directory.
            if not files_list:
                self.task_error.emit()
                return

            self.status_update.emit("************************* Tests Summary *************************\n\n")
            global progress_bar_step
            global progress_bar_counter
            progress_bar_counter = 0
            for file in files_list:
                if file.endswith('html'):
                    full_file_name = 'file://' + dir_name + '/' + file
                    driver.get(full_file_name)
                    table = driver.find_element(By.TAG_NAME, 'table')
                    if first_iteration:
                        headers = table.find_elements(By.TAG_NAME, 'th')
                        headers_list = [header.text for header in headers]
                        # Adding an extra column for the overall result
                        headers_list.append('Overall Result')
                        # Headers styling dict
                        headers_style = {
                            "font": Font(name='Arial', size=12, bold=True),
                            "alignment": Alignment(horizontal='center', vertical='center')
                        }
                        # Calling the function to write the headers name to the excel file.
                        write_data_excel(test_results_excel_file, [headers_list], headers_style)
                        # Setting the first iteration flag to false to write the headers only once in the file.
                        first_iteration = False
                    
                    # Getting tests data from the HTML file
                    rows = table.find_elements(By.TAG_NAME, 'tr')
                    table_data = []
                    for row in rows:
                        cells = row.find_elements(By.TAG_NAME, 'td')
                        cell_data = [cell.text for cell in cells]
                        if cell_data:
                            table_data.append(cell_data)
                    # Data Styling dict
                    data_style = {
                        "font": Font(name='Arial', size=12, bold=False),
                        "alignment": Alignment(horizontal='center', vertical='center')
                    }
                    # Removing the last row in the HTML (contains the test result summary[pass or fail])
                    overall_test_result = table_data.pop()
                    # Appending the test result to the first row in the new column (overall results)
                    table_data[0].append(overall_test_result[0])
                    # Sending signals to update status on the GUI and update the progress bar
                    self.status_update.emit(f"*************Test File: {file}*************\n")
                    self.status_update.emit(f"{str(overall_test_result)}\n")
                    write_data_excel(test_results_excel_file, table_data, data_style)
                    progress_bar_counter += progress_bar_step
                    self.progress_updated.emit(int(progress_bar_counter))
            
            # Applying red highlight to failed tests and green highlight to passed tests
            passed_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
            failed_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            passed_cases = "Test Result : PASSED"
            failed_cases = "Test Result : FAILED"
            format_cells_with_values(test_results_excel_file, passed_cases, fill=passed_fill)
            format_cells_with_values(test_results_excel_file, failed_cases, fill=failed_fill)
            # Sending a signal that the task is completed.
            self.task_completed.emit()

        except Exception as e:
            print(e)
            self.task_error.emit()



class BackEndClass(QtWidgets.QWidget, Ui_MainWindow):
    progress_updated = pyqtSignal(int)  # Signal to update progress
    task_completed = pyqtSignal()       # Signal to notify task completion
    status_update = pyqtSignal(str)     # Signal to receive status updates
    task_error = pyqtSignal()           # Signal when error occurs

    def __init__(self):
        QtWidgets.QWidget.__init__(self)
        self.setupUi(MainWindow)
        self.browse_btn.clicked.connect(self.browse_function)
        self.start_btn.clicked.connect(self.start_threading)
        self.clear_btn.clicked.connect(self.clear_logs_function)

    def browse_function(self):
        self.browse_label.clear()
        try:
            global dir_name
            dir_name = QFileDialog.getExistingDirectory()
            global files_list
            files_list = os.listdir(dir_name)
            self.browse_label.setText(dir_name)
            files_list = [file for file in files_list if file.endswith('html')]
            if not files_list:
                QMessageBox.about(self, "Message", "Directory Doesn't Contain HTML files!")
                self.num_files_label.setText("No HTML files in this folder")
            global number_of_files
            number_of_files = len(files_list)
            global progress_bar_step
            progress_bar_step = 100/number_of_files
            self.num_files_label.setText(f"Number of HTML files in selected folder is {number_of_files}")
        except Exception as e:
            print(e)
            QMessageBox.about(self, "Message", "Please Select a valid Directory with HTML files.")

    def start_threading(self):
        # Create a worker instance and connect signals
        self.worker = Worker(parent=self)
        self.worker.progress_updated.connect(self.update_progress)
        self.worker.status_update.connect(self.update_status)
        self.worker.task_completed.connect(self.processing_complete)
        self.worker.task_error.connect(self.processing_error)

        # Create and start the thread
        self.thread = QThread()
        self.worker.moveToThread(self.thread)
        self.thread.started.connect(self.worker.run)
        self.thread.start()

    def update_progress(self, value):
        self.progressBar.setValue(value)

    def update_status(self, message):
        self.status_textEdit.append(message)  

    def processing_complete(self):
        self.status_textEdit.append("Excel File generated successfully !")
        self.thread.quit()

    def processing_error(self):
        self.thread.quit()
        QMessageBox.about(self, "Message", "No directory Selected! \n Please Select a Directory with HTML files.")
    
    def clear_logs_function(self):
        self.status_textEdit.clear()
        

if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = BackEndClass()
    MainWindow.show()
    sys.exit(app.exec_())
