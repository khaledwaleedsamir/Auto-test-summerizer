"""
**********************************************************************************
File: backend_final.py
Author: Khaled Waleed Samir Metwally
Date: 29/7/2024

Description:
This script is designed to process HTML test result files, extract relevant data,
and compile the results into an Excel file. The tool uses the PyQt5 library for 
the graphical user interface, selenium for web automation, and openpyxl for 
manipulating Excel files.

The script includes functionality for:
- Checking if an Excel file exists and creating one if it doesn't.
- Writing data to an Excel file.
- Applying custom styles to Excel sheets.
- Formatting cells based on specific search values.
- Multi-threaded processing to ensure the GUI remains responsive during long operations.

Classes:
- Worker: Handles the background processing of HTML files.
- BackEndClass: Manages the GUI interactions and starts the background thread.

Functions:
- check_if_excel_exists: Checks if an Excel file exists and returns an openpyxl Workbook object.
- is_sheet_empty: Checks if a specified sheet in an Excel file is empty.
- style_excel_sheet: Applies specified styling to the headers and data cells in an Excel sheet.
- write_data_excel: Writes data to an Excel file.
- format_cells_with_values: Formats cells in an Excel file that contain a specific search value.

Global Variables:
- dir_name: Holds the selected directory name.
- number_of_files: Contains the number of HTML files in the directory.
- progress_bar_counter: Holds the current progress bar counter.
- progress_bar_step: Defines the step of the progress bar based on the number of files to process.
- files_list: List containing the file names of the selected directory.
- options: Holds the Chrome driver options.
- driver: WebDriver instance for Google Chrome.
- test_results_excel_file: Holds the output Excel file name.

Usage:
1. Select a directory containing HTML test result files using the browse button.
2. Click the start button to process the files and generate an Excel file.
3. The status and progress will be updated in the GUI during processing.
4. The results will be saved in an Excel file named "Tests_Results.xlsx".

**********************************************************************************
"""

# Imports
import openpyxl.workbook
from front import Ui_MainWindow
from PyQt5 import QtWidgets
import sys
from PyQt5.QtWidgets import QMessageBox, QFileDialog
from PyQt5.QtCore import pyqtSignal, QObject, QThread
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os
import sys
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options

# Global Variables
global dir_name                        # Global variable to hold the selected directory name.
global number_of_files                 # Global variable that contains the number of html files in the directory.
number_of_files = 0                    # Initalizing the number of files global variable to 0 files
global progress_bar_counter            # Global variable holding the current progress bar counter
progress_bar_counter = 0               # Initalizing the progress bar counter to 0
global progress_bar_step               # Global variable for the step of the progress bar based on the number of files to process
global files_list                      # Global list of files that contains the file names of the selected directory.
options = Options()                    # Initalizing options variable to contain chrome driver options.
options.add_argument('--headless=new') # Headless option to run chrome in silent mode.
driver = webdriver.Chrome(options)     # Initalizing the WebDriver to google chrome with the created options.
test_results_excel_file = "Tests_Results.xlsx" # Global Variable holding the output excel file name


# Functions
def check_if_excel_exists(excel_file_name):
    """
    Checks if an Excel file exists, and returns an openpyxl Workbook object.

    If the specified Excel file exists, it loads and returns the workbook. 
    If the file does not exist, it creates a new workbook, saves it with the 
    specified filename, and returns the new workbook.

    Args:
        excel_file_name (str): The name or path of the Excel file to check.

    Returns:
        openpyxl.Workbook: The loaded or newly created workbook object.
    """
    if os.path.exists(excel_file_name):
        workbook = openpyxl.load_workbook(excel_file_name)
    else:
        workbook = openpyxl.Workbook()
        workbook.save(excel_file_name)
    return workbook


def is_sheet_empty(excel_file):
  """Checks if a specified sheet in an Excel file is empty.

  Args:
    excel_file: Path to the Excel file.

  Returns:
    True if the sheet is empty, False otherwise.
  """
  workbook = openpyxl.load_workbook(excel_file)
  sheet = workbook.active
  # Check if the sheet has any rows
  if sheet.max_row == 1 and sheet['A1'].value is None:
    return True
  else:
    return False

def style_excel_sheet(excel_file_name, header_styling=None, data_styling=None):
    """
    Applies specified styling to the headers and data cells in an Excel sheet.
    Contains some default styling options that will be used if header_styling or data_styling are not passed.
    
    Args:
        excel_file_name (str): The name or path of the Excel file to style.
        header_styling (dict): A dictionary of styling options for the header row.
        data_styling (dict): A dictionary of styling options for the data rows.
        
    Returns:
        None
    """
    # Defining the default styling options
    default_options = {
        "font": Font(name='Arial', size=14, bold=False),
        "alignment": Alignment(horizontal='center', vertical='center'),
        "fill": PatternFill(fill_type=None)
    }
    # check if there are any styling options passed to the function and combine them with default options
    header_styling = {**default_options, **header_styling}
    data_styling = {**default_options, **data_styling}
    # Load the workbook and select the active sheet
    workbook = check_if_excel_exists(excel_file_name)   
    sheet = workbook.active
     # Apply header styling
    for cell in sheet[1]:
        if 'font' in header_styling:
            cell.font = header_styling['font']
        if 'alignment' in header_styling:
            cell.alignment = header_styling['alignment']
        if 'fill' in header_styling:
            cell.fill = header_styling['fill']
     # Apply data styling
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if 'font' in data_styling:
                cell.font = data_styling['font']
            if 'alignment' in data_styling:
                cell.alignment = data_styling['alignment']
            if 'fill' in data_styling:
                cell.fill = data_styling['fill']
    
    # Autofitting the rows and columns the adjusted width and height formulas are obtained by trial and error and not accurate
    # Autofit columns
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
                    font_size = header_styling.get('font').size
    
            except:
                pass
        adjusted_width = max_length * font_size / 12 + 2
        sheet.column_dimensions[column].width = adjusted_width
    # Autofit rows
    for row in sheet.rows:
        max_height = 0
        for cell in row:
            try:
                if len(str(cell.value)) > max_height:
                    max_height = len(str(cell.value))
            except:
                pass
        adjusted_height = max_height * font_size / 12 + 2
        sheet.row_dimensions[row[0].row].height = adjusted_height
    
    workbook.save(excel_file_name)

def write_data_excel(excel_file_name, excel_data):
    """
    Writes data to an Excel file.

    Args:
        excel_file_name (str): The name or path of the Excel file to write data to.
        excel_data (list): A list of lists, where each sublist represents a row of data to be written to the Excel file.

    Returns:
        None
    """
    # Load the workbook and select the active sheet
    workbook = check_if_excel_exists(excel_file_name)
    sheet = workbook.active
    if is_sheet_empty(excel_file_name):
        next_row = sheet.max_row
    else:
        next_row = sheet.max_row + 1
    for row_data in excel_data:
        for col, data in enumerate(row_data, start=1):
            sheet.cell(row=next_row, column=col, value=data)
        next_row += 1
    workbook.save(excel_file_name)
    
    

def format_cells_with_values(excel_file_name, search_value, font=None, fill=None, alignment=None):
    """
    Formats cells in an Excel file that contain a specific search value with given styling options.

    This function loads an Excel file and iterates through all cells in the active sheet. 
    If a cell contains the specified search value, it applies the provided formatting options 
    (font, fill, and alignment) to that cell.

    Args:
        excel_file_name (str): The name or path of the Excel file to be formatted.
        search_value (any): The value to search for in the cells. Cells containing this value will be formatted.
        font (openpyxl.styles.Font, optional): The font to apply to the cells containing the search value. Default is None.
        fill (openpyxl.styles.PatternFill, optional): The fill pattern to apply to the cells containing the search value. Default is None.
        alignment (openpyxl.styles.Alignment, optional): The alignment to apply to the cells containing the search value. Default is None.

    Returns:
        None
    """
    # Load the workbook and select the active sheet
    workbook = openpyxl.load_workbook(excel_file_name)
    sheet = workbook.active
    
    # Iterate over all rows and columns in the sheet
    for row in sheet.iter_rows():
        for cell in row:
            # Check if the cell value matches the search value
            if cell.value == search_value:
                # Apply formatting if value matches
                if font:
                    cell.font = font
                if fill:
                    cell.fill = fill
                if alignment:
                    cell.alignment = alignment
    
    # Save the workbook
    workbook.save(excel_file_name)


# Classes
class Worker(QObject):
    # Signals to control the GUI using threading
    progress_updated = pyqtSignal(int)  # Signal to update progress
    status_update = pyqtSignal(str)     # Signal to send status updates
    task_completed = pyqtSignal()       # Signal of task completion
    task_error = pyqtSignal()           # Signal of error

    def __init__(self, parent=None):
        super().__init__()
        self.parent_widget = parent

    def run(self):
        try:
            # Remove any previous excel files to start with a new one
            if os.path.exists(test_results_excel_file):
                os.remove(test_results_excel_file)
            # Variable to check if this is the first iteration.
            first_iteration = True
            # Check if there are HTML files in the directory.
            if not files_list:
                self.task_error.emit()
                return

            self.status_update.emit("************************* Tests Summary *************************\n\n")
            global progress_bar_step
            global progress_bar_counter
            progress_bar_counter = 0
            for file in files_list:
                if file.endswith('html'):
                    full_file_name = 'file://' + dir_name + '/' + file
                    driver.get(full_file_name)
                    table = driver.find_element(By.TAG_NAME, 'table')
                    if first_iteration:
                        headers = table.find_elements(By.TAG_NAME, 'th')
                        headers_list = [header.text for header in headers]
                        print(headers_list)
                        # Adding an extra column for the overall result
                        headers_list.append('Overall Result')
                        # Headers styling dict
                        # Calling the function to write the headers name to the excel file.
                        write_data_excel(test_results_excel_file, [headers_list])
                        # Setting the first iteration flag to false to write the headers only once in the file.
                        first_iteration = False
                    
                    # Getting tests data from the HTML file
                    rows = table.find_elements(By.TAG_NAME, 'tr')
                    table_data = []
                    for row in rows:
                        cells = row.find_elements(By.TAG_NAME, 'td')
                        cell_data = [cell.text for cell in cells]
                        if cell_data:
                            table_data.append(cell_data)
                    # Removing the last row in the HTML (contains the test result summary[pass or fail])
                    overall_test_result = table_data.pop()
                    # Appending the test result to the first row in the new column (overall results)
                    table_data[0].append(overall_test_result[0])
                    # Sending signals to update status on the GUI and update the progress bar
                    self.status_update.emit(f"*************Test File: {file}*************\n")
                    self.status_update.emit(f"{str(overall_test_result)}\n")
                    print(table_data) # print for testing
                    # Write the data of the current HTML file to the excel file
                    write_data_excel(test_results_excel_file, table_data)
                    progress_bar_counter += progress_bar_step
                    self.progress_updated.emit(int(progress_bar_counter))

            ####################################### Styling The Excel Sheet #######################################
            # Headers Styling dict
            headers_style = {
                    "font": Font(name='Arial', size=12, bold=True),
                    "alignment": Alignment(horizontal='center', vertical='center'),
                    "fill": PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
                }
            # Data Styling dict
            data_style = {
                "font": Font(name='Arial', size=12, bold=False),
                "alignment": Alignment(horizontal='center', vertical='center')
            }
            # Applying style to the Excel sheet
            style_excel_sheet(test_results_excel_file, headers_style, data_style)
            # Applying red highlight to failed tests and green highlight to passed tests
            passed_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
            failed_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            passed_cases = "Test Result : PASSED"
            failed_cases = "Test Result : FAILED"
            format_cells_with_values(test_results_excel_file, passed_cases, fill=passed_fill)
            format_cells_with_values(test_results_excel_file, failed_cases, fill=failed_fill)
            # Sending a signal that the task is completed.
            self.task_completed.emit()

        except Exception as e:
            print(e)
            self.task_error.emit()



class BackEndClass(QtWidgets.QWidget, Ui_MainWindow):
    progress_updated = pyqtSignal(int)  # Signal to update progress
    task_completed = pyqtSignal()       # Signal to notify task completion
    status_update = pyqtSignal(str)     # Signal to receive status updates
    task_error = pyqtSignal()           # Signal when error occurs

    def __init__(self):
        QtWidgets.QWidget.__init__(self)
        self.setupUi(MainWindow)
        self.browse_btn.clicked.connect(self.browse_function)
        self.start_btn.clicked.connect(self.start_threading)
        self.clear_btn.clicked.connect(self.clear_logs_function)

    def browse_function(self):
        self.browse_label.clear()
        try:
            global dir_name
            dir_name = QFileDialog.getExistingDirectory()
            global files_list
            files_list = os.listdir(dir_name)
            self.browse_label.setText(dir_name)
            files_list = [file for file in files_list if file.endswith('html')]
            if not files_list:
                QMessageBox.about(self, "Message", "Directory Doesn't Contain HTML files!")
                self.num_files_label.setText("No HTML files in this folder")
            global number_of_files
            number_of_files = len(files_list)
            global progress_bar_step
            progress_bar_step = 100/number_of_files
            self.num_files_label.setText(f"Number of HTML files in selected folder is {number_of_files}")
        except Exception as e:
            print(e)
            QMessageBox.about(self, "Message", "Please Select a valid Directory with HTML files.")

    def start_threading(self):
        # Create a worker instance and connect signals
        self.worker = Worker(parent=self)
        self.worker.progress_updated.connect(self.update_progress)
        self.worker.status_update.connect(self.update_status)
        self.worker.task_completed.connect(self.processing_complete)
        self.worker.task_error.connect(self.processing_error)

        # Create and start the thread
        self.thread = QThread()
        self.worker.moveToThread(self.thread)
        self.thread.started.connect(self.worker.run)
        self.thread.start()

    def update_progress(self, value):
        # Disable the button to prevent multiple runs
        self.start_btn.setEnabled(False)
        self.progressBar.setValue(value)

    def update_status(self, message):
        self.status_textEdit.append(message)  

    def processing_complete(self):
        self.status_textEdit.append("Excel File generated successfully !")
        # Re-enable the button after the processing is done
        self.start_btn.setEnabled(True)
        self.thread.quit()

    def processing_error(self):
        self.thread.quit()
        if number_of_files == 0:
            QMessageBox.about(self, "Message", "No directory Selected! \n Please Select a Directory with HTML files.")
        else:
            QMessageBox.about(self, "Message", "Empty HTML FILE")


    def clear_logs_function(self):
        self.status_textEdit.clear()
        

if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = BackEndClass()
    MainWindow.show()
    sys.exit(app.exec_())
