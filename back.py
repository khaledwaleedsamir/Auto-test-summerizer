# Imports
import openpyxl.workbook
from front import Ui_MainWindow
from PyQt5 import QtWidgets
import sys
from PyQt5.QtWidgets import QMessageBox, QFileDialog, QProgressBar
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os
import time
import sys
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver import ActionChains
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options
import threading

# Global Variables
global dir_name                        # Global variable to hold the selected directory name.
global number_of_files                 # Global variable that contains the number of html files in the directory.
number_of_files = 0                    # Initalizing the number of files global variable to 0 files
global progress_bar_counter            # Global variable holding the current progress bar counter
progress_bar_counter = 0               # Initalizing the progress bar counter to 0
global progress_bar_step               # Global variable for the step of the progress bar based on the number of files to process
global files_list                      # Global list of files that contains the file names of the selected directory.
options = Options()                    # Initalizing options variable to contain chrome driver options.
options.add_argument('--headless=new') # Headless option to run chrome in silent mode.
driver = webdriver.Chrome(options)     # Initalizing the WebDriver to google chrome with the created options.
test_results_excel_file = "Tests_Results.xlsx" # Global Variable holding the output excel file name


# Check if Excel file exists
def check_if_excel_exists(excel_file_name):
    if os.path.exists(excel_file_name):
        workbook = openpyxl.load_workbook(excel_file_name)
    else:
        workbook = openpyxl.Workbook()
        workbook.save(excel_file_name)
    return workbook

def is_sheet_empty(excel_file):
  """Checks if a specified sheet in an Excel file is empty.

  Args:
    excel_file: Path to the Excel file.

  Returns:
    True if the sheet is empty, False otherwise.
  """
  workbook = openpyxl.load_workbook(excel_file)
  sheet = workbook.active
  # Check if the sheet has any rows
  if sheet.max_row == 1 and sheet['A1'].value is None:
    return True
  else:
    return False


# Function for Writing data to Excel file
def write_data_excel(excel_file_name, excel_data, styling_options):
        # Defining the default styling options
        default_options = {
            "font": Font(name='Arial', size=14, bold=False),
            "alignment": Alignment(horizontal='center', vertical='center')
        }
        # check if there are any styling options passed to the function and combine them with default options
        styling_options = {**default_options, **styling_options}
        workbook = check_if_excel_exists(excel_file_name)   
        # Load the workbook and select the active sheet
        sheet = workbook.active
        if is_sheet_empty(excel_file_name):
            next_row = sheet.max_row
        else:
            next_row = sheet.max_row + 1
        for row_data in excel_data:
            for col, data in enumerate(row_data, start=1):
                cell = sheet.cell(row=next_row, column=col, value=data)
                cell.font = styling_options.get('font')
                cell.alignment = styling_options.get('alignment')
            next_row += 1
        
        # Autofitting the rows and columns the adjusted width and height formulas are obtained by trial and error and not accurate
        # Autofit columns
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get column letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                        font_size = styling_options.get('font').size
        
                except:
                    pass
            adjusted_width = max_length * font_size / 12 + 2
            sheet.column_dimensions[column].width = adjusted_width
        # Autofit rows
        for row in sheet.rows:
            max_height = 0
            for cell in row:
                try:
                    if len(str(cell.value)) > max_height:
                        max_height = len(str(cell.value))
                except:
                    pass
            adjusted_height = max_height * font_size / 12 + 2
            sheet.row_dimensions[row[0].row].height = adjusted_height
        
        workbook.save(excel_file_name)

def format_cells_with_values(excel_file_name, search_value, font=None, fill=None, alignment=None):
    # Load the workbook and select the active sheet
    workbook = openpyxl.load_workbook(excel_file_name)
    sheet = workbook.active
    
    # Iterate over all rows and columns in the sheet
    for row in sheet.iter_rows():
        for cell in row:
            # Check if the cell value matches the search value
            if cell.value == search_value:
                # Apply formatting if value matches
                if font:
                    cell.font = font
                if fill:
                    cell.fill = fill
                if alignment:
                    cell.alignment = alignment
    
    # Save the workbook
    workbook.save(excel_file_name)


class BackEndClass(QtWidgets.QWidget, Ui_MainWindow):

    def __init__(self):
        QtWidgets.QWidget.__init__(self)
        self.setupUi(MainWindow)
        self.browse_btn.clicked.connect(self.browse_function)
        self.start_btn.clicked.connect(self.start_function)

    def browse_function(self):
        self.browse_label.clear()
        try:
            # Getting the list of files in the selected directory
            global dir_name
            dir_name = QFileDialog.getExistingDirectory()
            global files_list
            files_list = os.listdir(dir_name)
            # updating the browse label with the selected directory
            self.browse_label.setText(dir_name)
            for file in files_list:
                if file.endswith('html'):
                    pass
                else:
                    files_list.remove(file)
            if not files_list:
                QMessageBox.about(self, "Message", "Please Select a valid Directory with HTML files.")
            global number_of_files
            number_of_files = len(files_list)
            self.num_files_label.setText(f"Number of HTML files in selected folder is {number_of_files}")
        except Exception as e:
            print(e)
            QMessageBox.about(self, "Message", "Please Select a valid Directory with HTML files.")
        

    def start_function(self):
        try:
            # Remove any previous excel files to start with a new one
            if os.path.exists(test_results_excel_file):
                os.remove(test_results_excel_file)
            # Clear the text edit widget
            self.status_textEdit.clear()
            # Boolean variable to check if this is the first iteration
            first_iteration = True
            # Check if the files_list contains HTML files 
            if not files_list:
                self.status_textEdit.insertPlainText("No Available HTML files in this directory! \n")
                QMessageBox.about(self, "Message", "Please Select a valid Directory with HTML files.")
            # HTML files Available start the processing on them
            if files_list:
                self.status_textEdit.insertPlainText("************************* Tests Summary ************************* \n \n",)
            # Loop over all the html files to extract important data from tests results
            for file in files_list:
                # Only look at html files
                if file.endswith('html'):
                    # Get full file name -> example: 'file://full_path/file_name'
                    full_file_name = 'file://' + dir_name + '/' + file
                    # open the file using chrome's driver
                    driver.get(full_file_name)
                    # finding the table, table headers and table entries by tag name
                    table = driver.find_element(By.TAG_NAME, 'table')
                    # Check if this is the first iteration to write the headers once in the excel file
                    if first_iteration:
                        headers = table.find_elements(By.TAG_NAME, 'th')
                        headers_list = [header.text for header in headers]
                        headers_list.append('Overall Result')
                        print(headers_list) # print for testing
                        headers_style = {
                        "font": Font(name='Arial', size=12, bold=True),
                        "alignment": Alignment(horizontal='center', vertical='center')
                        }
                        write_data_excel(test_results_excel_file, [headers_list], headers_style)
                        first_iteration = False

                    rows = table.find_elements(By.TAG_NAME, 'tr')
                    table_data = []
                    for row in rows:
                        cells = row.find_elements(By.TAG_NAME, 'td')
                        cell_data = [cell.text for cell in cells]
                        if cell_data:
                            table_data.append(cell_data)
                    print(table_data) # print for testing
                    data_style = {
                        "font": Font(name='Arial', size=12, bold=False),
                        "alignment": Alignment(horizontal='center', vertical='center')
                    }
                    self.status_textEdit.insertPlainText(f"*************Test File: {file}************* \n")
                    # obtain the overall test result as it is the last entry
                    overall_test_result = table_data.pop()
                    table_data[0].append(overall_test_result[0])  # Append the test result to the first row
                    self.status_textEdit.insertPlainText(f"{str(overall_test_result)} \n")
                    write_data_excel(test_results_excel_file, table_data, data_style)
            
            # Adjusting the excel file with the passed as green and failed as red highlight
            passed_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
            failed_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            passed_cases = "Test Result : PASSED"
            failed_cases = "Test Result : FAILED"
            format_cells_with_values(test_results_excel_file, passed_cases, fill=passed_fill)
            format_cells_with_values(test_results_excel_file, failed_cases, fill=failed_fill)

        except Exception as e:
            print(e)
            QMessageBox.about(self, "Message", "Please Select a valid Directory with HTML files.")


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = BackEndClass()
    MainWindow.show()
    sys.exit(app.exec_())
